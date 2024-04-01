import configparser
from flask import Flask, Response, request, send_file
from flask_cors import CORS
import psycopg2
import websocket
import json
import openpyxl
from openpyxl.styles import Border, Alignment, Side
from datetime import datetime


config = configparser.ConfigParser()
config.read('conf.ini')

api_url = config.get('server', 'server_url')
api_user = config.get('server', 'server_user')
api_password = config.get('server', 'server_password')

db_host = config.get('db', 'db_host')
db_dbname = config.get('db', 'db_name')
db_user = config.get('db', 'db_user')
db_password = config.get('db', 'db_password')

vega_si_off = "040X"

app = Flask(__name__)
cors = CORS(app, resources={r'*': {'origins': '*'}}, supports_credentials=True, cors_allowed_origins="*")


@app.route('/reports/td', methods=['GET'])
def print_report_td():
    border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'))
    alignment = Alignment(horizontal='center', vertical='center')

    data_logging = {
        'api_con_st': '',
        'api_auth_st': '',
        'db_con_st': '',
        'db_query_st': '',
        'si_sw': '',
        'si_sw_st': ''
    }
    device_td_id = request.args.get('device_td_id')
    time_from = request.args.get('time_from')
    time_to = request.args.get('time_to')

    wb = openpyxl.load_workbook('/opt/iot-vega-programs/template_report_td.xlsx')
    ws = wb.worksheets[0]
    dt_time_from = datetime.fromtimestamp(int(time_from) / 1000).strftime('%d.%m.%Y %H:%M:%S')
    dt_time_to = datetime.fromtimestamp(int(time_to) / 1000).strftime('%d.%m.%Y %H:%M:%S')
    # print(dt_time_from)
    # print(dt_time_to)

    try:
        conn = psycopg2.connect(dbname=db_dbname, user=db_user, password=db_password, host=db_host)
        cursor = conn.cursor()
        data_logging.update({'db_con_st': True})

        try:
            # Get Inventory to Device TD
            cursor.execute("""
                SELECT d.devname AS device_name,
                       da.value::json->>'address_level_1' AS obj_address,
                       da.value::json->>'name_level_1' AS obj_name,
                       da.value::json->>'level_2' AS placement
                FROM devices AS d
                LEFT JOIN deviceattributes AS da ON d.deveui = da.deveui
                WHERE d.deveui = '{0}'
                  AND da.attribute = 'other_info_1';
            """.format(device_td_id))
            record = cursor.fetchone()
            # print(record)
            device_td_name = record[0]
            obj_address = record[1]
            obj_name = record[2]
            placement = record[3]

            # Get Data to Device TD
            cursor.execute("""
                SELECT rd.time AS time,
                ROUND(temperature(CONCAT(SUBSTRING(encode(rd.data, 'hex'),17,2), SUBSTRING(encode(rd.data, 'hex'),15,2))), 1) AS t_cur,
                ROUND(('x' || SUBSTRING(encode(rd.data, 'hex'),19,2))::bit(8)::int, 1) as t_min,
                ROUND(('x' || SUBSTRING(encode(rd.data, 'hex'),21,2))::bit(8)::int, 1) as t_max,
                ('x' || SUBSTRING(encode(rd.data, 'hex'),3,2))::bit(8)::int as t_bat
                FROM rawdata AS rd
                WHERE rd.deveui = '{0}'
                AND (rd.time >= {1} AND rd.time <= {2})
                AND SUBSTRING(encode(rd.data, 'hex'),1,2) = '01';
            """.format(device_td_id, time_from, time_to))
            records = cursor.fetchall()
            data_logging.update({'db_query_st': True})

            ws['C4'] = dt_time_from
            ws['C5'] = dt_time_to
            ws['C6'] = device_td_name
            ws['C7'] = device_td_id
            ws['C8'] = obj_address
            ws['C9'] = obj_name
            ws['C10'] = placement

            row = 13
            for record in records:
                u_time = datetime.fromtimestamp(int(record[0]) / 1000).strftime('%d.%m.%Y %H:%M:%S')
                t_cur = record[1]
                t_min = record[2]
                t_max = record[3]
                td_battery = record[4]

                ws.cell(row=row, column=1).value = u_time
                ws.cell(row=row, column=1).alignment = alignment
                ws.cell(row=row, column=1).border = border

                ws.cell(row=row, column=2).value = t_cur
                ws.cell(row=row, column=2).alignment = alignment
                ws.cell(row=row, column=2).border = border

                ws.cell(row=row, column=3).value = t_min
                ws.cell(row=row, column=3).alignment = alignment
                ws.cell(row=row, column=3).border = border

                ws.cell(row=row, column=4).value = t_max
                ws.cell(row=row, column=4).alignment = alignment
                ws.cell(row=row, column=4).border = border

                ws.cell(row=row, column=5).value = td_battery
                ws.cell(row=row, column=5).alignment = alignment
                ws.cell(row=row, column=5).border = border
                row += 1

        except psycopg2.Error:
            data_logging.update({'db_query_st': False})

        cursor.close()
        conn.close()

    except psycopg2.Error:
        data_logging.update({'db_con_st': False})

    print(data_logging)
    wb.save('/opt/iot-vega-programs/Отчет по ТД.xlsx')
    return send_file('Отчет по ТД.xlsx', as_attachment=True)


@app.route('/si', methods=['GET'])
def deactivate_si():
    data_logging = {
        'api_con_st': '',
        'api_auth_st': '',
        'db_con_st': '',
        'db_query_st': '',
        'si_sw': '',
        'si_sw_st': ''
    }
    si_id = request.args.get('si_id')
    si_port = request.args.get('si_port')

    try:
        ws = websocket.WebSocket()
        ws.connect(vega_api_url)
        data_logging.update({'api_con_st': True})

        message_auth_req = {"cmd": "auth_req", "login": vega_api_user, "password": vega_api_password}
        ws.send(json.dumps(message_auth_req))
        srv_answer_auth_req = json.loads(ws.recv())
        auth_req_status = srv_answer_auth_req["status"]
        auth_req_token = srv_answer_auth_req["token"]
        data_logging.update({'api_auth_st': auth_req_status})

        try:
            conn = psycopg2.connect(dbname=db_dbname, user=db_user, password=db_password, host=db_host)
            cursor = conn.cursor()
            data_logging.update({'db_con_st': True})

            try:
                cursor.execute("""
                    SELECT CAST(SUBSTRING(encode(rd.data, 'hex'),10,1) AS INT) as data
                    FROM rawdata AS rd
                    LEFT JOIN
                    (
                        SELECT MAX(rd.time) AS time, rd.deveui AS deveui
                        FROM rawdata AS rd
                        WHERE rd.deveui = '{0}'
                        AND rd.port = 2
                        AND SUBSTRING(encode(rd.data, 'hex'),1,2) = '05'
                        AND SUBSTRING(encode(rd.data, 'hex'),8,1) = '{1}'
                        GROUP BY rd.deveui
                    ) table_1 ON rd.deveui = table_1.deveui
                    WHERE rd.time = table_1.time
                    AND rd.port = 2
                    AND SUBSTRING(encode(rd.data, 'hex'),1,2) = '05'
                    AND SUBSTRING(encode(rd.data, 'hex'),8,1) = '{1}';
                """.format(si_id, si_port))
                records = cursor.fetchone()
                data_logging.update({'db_query_st': True})

                if records:
                    if records[0] == 1:
                        data_logging.update({'si_sw': True})
                        si_data = vega_si_off.replace('X', si_port)

                        message_send_data_req = {"cmd": "send_data_req", "data_list": [{"devEui": si_id, "data": si_data, "port": 2, "ack": True}]}
                        ws.send(json.dumps(message_send_data_req))
                        srv_answer_data_req = json.loads(ws.recv())
                        send_data_req_status = srv_answer_data_req["status"]
                        data_logging.update({'si_sw_st': send_data_req_status})
                    else:
                        data_logging.update({'si_sw': False})

            except psycopg2.Error:
                data_logging.update({'db_query_st': False})

            cursor.close()
            conn.close()

        except psycopg2.Error:
            data_logging.update({'db_con_st': False})

        message_close_auth = {"cmd": "close_auth_req", "token": auth_req_token}
        ws.send(json.dumps(message_close_auth))
        ws.recv()
        ws.close()

    except ConnectionRefusedError:
        data_logging.update({'api_con_st': False})

    print(data_logging)
    return Response(status=204)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
